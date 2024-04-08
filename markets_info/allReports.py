import time
import os
import pandas as pd
import numpy as np
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from pandas.tseries.offsets import MonthEnd
from sql_query import fetch_data
from dotenv import load_dotenv
load_dotenv()

current_dir = os.path.dirname(os.path.abspath(__file__))

# Проверяем, находимся ли мы уже в папке markets_info
if os.path.basename(current_dir) == 'markets_info':
    file_path = os.path.join(current_dir, 'allReports.xlsx')
else:
    # Если нет, ищем папку markets_info внутри текущей директории
    file_path = os.path.join(current_dir, 'markets_info', 'allReports.xlsx')


def apply_style_to_sheet(sheet):
    sheet.freeze_panes = 'B4'
    print('шрифт, выравнивание')
    for row in range(2, sheet.max_row + 1):
        for col in range(2, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            if row == 2 or row == 3:
                cell.font = Font(bold=True)
                cell.number_format = '#,##0.00 ₽' if row == 3 else '#,##0'

    for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
        max_length = max(len(str(cell.value)) for cell in col if cell.value) + 2
        sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length * 1.2

    print('ширина столбцов')
    # Группировка столбцов по дням месяца, оставляя видимыми суммирующие столбцы месяцев
    start_group_col = None
    for col in range(2, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value and " " in cell_value:  # Суммирующий столбец
            if start_group_col:
                sheet.column_dimensions.group(get_column_letter(start_group_col), get_column_letter(col - 1),
                                              hidden=True)
                start_group_col = None
        else:
            if start_group_col is None:
                start_group_col = col

def aggregate(df):
    df['order_date'] = pd.to_datetime(df['order_date'])
    df['year'] = df['order_date'].dt.year  # Вычисляем год здесь
    df['month_name'] = df['order_date'].dt.strftime('%B %Y')
    df['date_label'] = df['order_date'].dt.strftime('%d-%m-%Y')
    df = df.sort_values(by=['order_date', 'marketplace'])  # Сортировка данных по дате

    daily_sales = df.groupby(['year', 'marketplace', 'date_label', 'products_model']).agg(
        {'bought_pc': 'sum',
         'bought_price': 'sum'}
    ).reset_index()

    monthly_sales = df.groupby(['year', 'marketplace', 'month_name', 'products_model']).agg(
        {'bought_pc': 'sum',
         'bought_price': 'sum'}
    ).reset_index()

    daily_total = df.groupby(['year', 'marketplace', 'date_label']).agg(
        {'bought_pc': 'sum',
         'bought_price': 'sum'}
    ).reset_index()

    monthly_total = df.groupby(['year', 'marketplace', 'month_name']).agg(
        {'bought_pc': 'sum',
         'bought_price': 'sum'}
    ).reset_index()

    return df, daily_sales, monthly_sales, daily_total, monthly_total

def create_pattern(df, empty_tables):
    for (year, marketplace), group_df in df.groupby(['year', 'marketplace']):
        start_date = pd.to_datetime(f'{year}-01-01')
        end_date = pd.to_datetime(f'{year}-12-31')

        date_range = pd.date_range(start=start_date, end=end_date, freq='D')
        final_dates = []

        for date in date_range:
            final_dates.append(date.strftime("%d-%m-%Y"))
            if date == date + MonthEnd(0) - pd.offsets.Day(0):
                final_dates.append(date.strftime("%B %Y"))

        # Индексные метки для строк, начиная с 'Дата', 'Общее кол-во', 'Общая сумма'
        df_index = ['Кол-во', 'Сумма']
        unique_products = np.unique(group_df['products_model'].dropna())
        unique_products = [product for product in unique_products if product != '']
        unique_products = sorted(unique_products)  # Сортировка по алфавиту
        df_index.extend(unique_products)

        # Создаем DataFrame
        new_df = pd.DataFrame(index=df_index, columns=final_dates).fillna('')
        # Добавляем название индекса 'Дата'
        new_df.index.name = 'Дата'
        empty_tables[(year, marketplace)] = new_df

    return empty_tables

def for_excel(pattern, dict_for_excel):
    for (year, marketplace), schema in pattern.items():
        print(year, marketplace)
        schema.loc[:, schema.columns.str.contains(' ')] = 0
        # Преобразование daily_sales и daily_total для текущего year и marketplace
        daily_data = daily_sales[(daily_sales['year'] == year) & (daily_sales['marketplace'] == marketplace)]
        daily_total_data = daily_total[(daily_total['year'] == year) & (daily_total['marketplace'] == marketplace)]

        # Объединение daily_data и daily_total_data для заполнения 'Общее кол-во' и 'Общая сумма' за день
        for index, row in daily_data.iterrows():
            date_label = row['date_label']
            product_model = row['products_model']
            schema.loc[product_model, date_label] = row['bought_pc']

        for index, row in daily_total_data.iterrows():
            date_label = row['date_label']
            schema.loc['Кол-во', date_label] = row['bought_pc']
            schema.loc['Сумма', date_label] = row['bought_price']

        # Аналогичное заполнение для monthly_sales и monthly_total
        monthly_data = monthly_sales[(monthly_sales['year'] == year) & (monthly_sales['marketplace'] == marketplace)]
        monthly_total_data = monthly_total[
            (monthly_total['year'] == year) & (monthly_total['marketplace'] == marketplace)]

        for index, row in monthly_data.iterrows():
            month_name = row['month_name']
            product_model = row['products_model']
            schema.loc[product_model, month_name] = row['bought_pc']

        for index, row in monthly_total_data.iterrows():
            month_name = row['month_name']
            schema.loc['Кол-во', month_name] = row['bought_pc']
            schema.loc['Сумма', month_name] = row['bought_price']

        dict_for_excel[(year, marketplace)] = schema
    return dict_for_excel

start_time = time.time()

user = os.getenv('USER')
password = os.getenv('PASSWORD')
host = os.getenv('HOST')
db_name = os.getenv('DB_NAME')

database_uri = f"mysql+pymysql://{user}:{password}@{host}:3306/{db_name}"
# database_uri = f"mysql+pymysql://{user}:{password}@{db_name}?unix_socket=/var/lib/mysql/mysql.sock" #для запуска на сервере

engine = create_engine(database_uri)
print('создан двигатель бд')


if not os.path.exists(file_path):
    print(f"Файл {file_path} не найден. Создание файла")
    last_date = '2019-01-01'
    df = fetch_data(engine, last_date)
    print('получен датафрейм с нуля')

    df, daily_sales, monthly_sales, daily_total, monthly_total = aggregate(df)
    print('получены агрегаты')

    empty_tables = {}
    print('создание группировки')
    pattern = create_pattern(df, empty_tables)

    dict_for_excel = {}
    print('создание данных для excel')
    data_for_excel = for_excel(pattern, dict_for_excel)

    end_time = time.time()
    elapsed_time = end_time - start_time
    minutes, seconds = divmod(elapsed_time, 60)
    print(f"Время выполнения: {int(minutes)} мин {int(seconds)} сек.")

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        for (year, marketplace), empty_table in data_for_excel.items():
            print("лист", year, marketplace)
            sheet_name = f'{marketplace} {year}'
            empty_table.to_excel(writer, sheet_name=sheet_name[:31], index=True)
            workbook = writer.book
            worksheet = workbook[sheet_name]
            apply_style_to_sheet(worksheet)

    print('Файл Excel создан и отформатирован')

else:
    print(f"Файл {file_path} найден. Дополнение файла")

    workbook = load_workbook(file_path)
    sheets = workbook.sheetnames
    last_year = sheets[-1].split(' ')[-1]
    last_date = f'{last_year}-01-01'
    print(last_year)
    df = fetch_data(engine, last_date)
    print('получен датафрейм (для дополнения)')

    df, daily_sales, monthly_sales, daily_total, monthly_total = aggregate(df)
    print('получены агрегаты')

    empty_tables = {}
    print('создание группировки')
    pattern = create_pattern(df, empty_tables)

    dict_for_excel = {}
    print('создание данных для excel')
    data_for_excel = for_excel(pattern, dict_for_excel)

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        book = writer.book
        existing_sheets = book.sheetnames
        for (year, marketplace), df in data_for_excel.items():
            sheet_name = f'{marketplace} {year}'[:31]
            print("Обработка листа:", sheet_name)
            # Если лист с таким именем уже существует, удаляем его
            if sheet_name in existing_sheets:
                std = book[sheet_name]
                book.remove(std)
            # Создаем лист заново и записываем обновленные данные
            df.to_excel(writer, sheet_name=sheet_name, index=True)
            worksheet = book[sheet_name]
            apply_style_to_sheet(worksheet)

end_time = time.time()
elapsed_time = end_time - start_time
minutes, seconds = divmod(elapsed_time, 60)
print(f"Время выполнения: {int(minutes)} мин {int(seconds)} сек.")

